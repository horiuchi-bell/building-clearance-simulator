#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OIRANシュミレーターExcelファイルの構造分析スクリプト
"""

import openpyxl
import pandas as pd
import json

def analyze_excel_structure():
    """Excelファイルの構造を分析"""
    excel_file = "OIRANシュミレーター（修正）20231215.xlsx"
    
    try:
        # Excelファイルを読み込み
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        print("=== OIRAN シュミレーター構造分析 ===")
        print(f"ファイル名: {excel_file}")
        print(f"シート数: {len(workbook.sheetnames)}")
        print("\n=== シート一覧 ===")
        
        for i, sheet_name in enumerate(workbook.sheetnames):
            print(f"{i+1}. {sheet_name}")
        
        # メインシート「限界余裕測定図 片線」の分析
        target_sheet = "限界余裕測定図 片線"
        
        if target_sheet in workbook.sheetnames:
            print(f"\n=== {target_sheet} シート詳細分析 ===")
            sheet = workbook[target_sheet]
            
            # 使用されているセル範囲を確認
            print(f"使用範囲: {sheet.min_row}行{sheet.min_column}列 ～ {sheet.max_row}行{sheet.max_column}列")
            
            # 重要なセルの値を確認
            print("\n=== 重要なセル値 ===")
            print(f"B11 (曲線半径): {sheet['B11'].value}")
            print(f"D11 (カント): {sheet['D11'].value}")
            print(f"B14 (軌道中心からの距離): {sheet['B14'].value}")
            print(f"D14 (レール踏面からの高さ): {sheet['D14'].value}")
            
            # 入力セル周辺の構造を確認
            print("\n=== 入力セル周辺の構造 ===")
            for row in range(9, 17):
                for col in range(1, 7):  # A-F列
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        col_letter = openpyxl.utils.get_column_letter(col)
                        print(f"{col_letter}{row}: {cell.value}")
            
            # 数式を含むセルを探す
            print("\n=== 数式を含むセル（サンプル） ===")
            formula_cells = []
            for row in range(1, min(50, sheet.max_row + 1)):
                for col in range(1, min(20, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)
                    if hasattr(cell, 'formula') and cell.formula and cell.formula.startswith('='):
                        col_letter = openpyxl.utils.get_column_letter(col)
                        formula_cells.append({
                            'cell': f"{col_letter}{row}",
                            'formula': cell.formula,
                            'value': cell.value
                        })
            
            # 最初の20個の数式を表示
            for i, formula_info in enumerate(formula_cells[:20]):
                print(f"{formula_info['cell']}: {formula_info['formula']} = {formula_info['value']}")
            
            if len(formula_cells) > 20:
                print(f"...他 {len(formula_cells) - 20} 個の数式セル")
            
            # グラフ部分（右側）のデータを探す
            print("\n=== グラフ関連データ探索 ===")
            # 通常グラフデータは特定の列に集中している
            for col in range(8, min(25, sheet.max_column + 1)):  # H列以降
                col_letter = openpyxl.utils.get_column_letter(col)
                col_data = []
                for row in range(1, min(100, sheet.max_row + 1)):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        col_data.append((row, cell.value))
                
                if len(col_data) > 5:  # データが多い列を表示
                    print(f"\n{col_letter}列のデータ（最初の10行）:")
                    for row, value in col_data[:10]:
                        print(f"  {col_letter}{row}: {value}")
                    if len(col_data) > 10:
                        print(f"  ...他 {len(col_data) - 10} 行")
        
        else:
            print(f"シート '{target_sheet}' が見つかりません。")
            print("利用可能なシート:")
            for sheet_name in workbook.sheetnames:
                print(f"  - {sheet_name}")
        
        workbook.close()
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    analyze_excel_structure()