#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
建築限界数値データを抽出し、座標データを解析するスクリプト
"""

import openpyxl
import json
import math

def extract_building_limit_data():
    """建築限界数値データを抽出"""
    excel_file = "OIRANシュミレーター（修正）20231215.xlsx"
    
    try:
        # 数式と値両方を読み込み
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # 基本的な建築限界データを抽出
        basic_data_sheets = [
            "建築限界数値データ 交流",
            "建築限界数値データ 　非電化",
            "建築限界数値データ 　直流"
        ]
        
        building_limit_data = {}
        
        for sheet_name in basic_data_sheets:
            if sheet_name in wb.sheetnames:
                print(f"=== {sheet_name} ===")
                sheet = wb[sheet_name]
                
                # 座標データを抽出
                coordinates = []
                for row in range(1, min(100, sheet.max_row + 1)):
                    x_val = sheet.cell(row=row, column=1).value  # A列
                    y_val = sheet.cell(row=row, column=2).value  # B列
                    
                    if isinstance(x_val, (int, float)) and isinstance(y_val, (int, float)):
                        coordinates.append((x_val, y_val))
                
                building_limit_data[sheet_name] = coordinates
                print(f"座標データ数: {len(coordinates)}")
                
                # 最初の10個と最後の10個の座標を表示
                print("座標データ（最初の10個）:")
                for i, (x, y) in enumerate(coordinates[:10]):
                    print(f"  {i+1}: ({x}, {y})")
                
                if len(coordinates) > 10:
                    print("座標データ（最後の10個）:")
                    for i, (x, y) in enumerate(coordinates[-10:]):
                        print(f"  {len(coordinates)-10+i+1}: ({x}, {y})")
                
                print()
        
        # 建築限界の形状を分析
        print("=== 建築限界形状分析 ===")
        for sheet_name, coords in building_limit_data.items():
            if coords:
                x_coords = [x for x, y in coords]
                y_coords = [y for x, y in coords]
                
                print(f"{sheet_name}:")
                print(f"  X座標範囲: {min(x_coords):.1f} ～ {max(x_coords):.1f}")
                print(f"  Y座標範囲: {min(y_coords):.1f} ～ {max(y_coords):.1f}")
                print(f"  座標数: {len(coords)}")
                print()
        
        # 計算シートの分析
        print("=== 表示データ計算シート分析 ===")
        calc_sheet_name = "表示データ計算シート　片線"
        
        if calc_sheet_name in wb.sheetnames:
            sheet = wb[calc_sheet_name]
            
            # 重要な計算結果を確認
            print("計算結果（最初の20行）:")
            for row in range(1, 21):
                aa_val = sheet.cell(row=row, column=27).value  # AA列
                ab_val = sheet.cell(row=row, column=28).value  # AB列
                
                if aa_val is not None or ab_val is not None:
                    print(f"  行{row}: AA={aa_val}, AB={ab_val}")
        
        # JSONとして保存
        print("=== データ保存 ===")
        with open("building_limit_data.json", "w", encoding="utf-8") as f:
            # 座標データをJSON形式で保存
            json_data = {}
            for sheet_name, coords in building_limit_data.items():
                json_data[sheet_name] = {
                    "coordinates": coords,
                    "count": len(coords)
                }
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        print("building_limit_data.json に保存しました")
        
        wb.close()
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()

def analyze_calculation_formulas():
    """計算式の詳細解析"""
    print("\n=== 建築限界計算式の解析 ===")
    
    # PDFから取得した計算式を実装
    def calculate_expansion_width(radius):
        """曲線による拡大幅計算"""
        if radius == 0:
            return 0
        return 23100 / radius  # 一般限界
    
    def calculate_upper_expansion_width(radius):
        """上部限界拡大幅計算"""
        if radius == 0:
            return 0
        return 11550 / radius  # 架空電車線による電気運転区間の上部限界
    
    def calculate_cant_angle(cant, gauge=1067):
        """カント角度計算"""
        return math.atan(cant / gauge)
    
    def calculate_slack(radius):
        """スラック量計算"""
        if radius < 200:
            return 20
        elif radius < 240:
            return 15
        elif radius < 320:
            return 10
        elif radius <= 440:
            return 5
        else:
            return 0
    
    # 計算例
    print("計算例:")
    test_radius = 300
    test_cant = 45
    
    expansion = calculate_expansion_width(test_radius)
    upper_expansion = calculate_upper_expansion_width(test_radius)
    cant_angle = calculate_cant_angle(test_cant)
    slack = calculate_slack(test_radius)
    
    print(f"曲線半径: {test_radius}m")
    print(f"カント: {test_cant}mm")
    print(f"拡大幅: {expansion:.1f}mm")
    print(f"上部拡大幅: {upper_expansion:.1f}mm")
    print(f"カント角度: {cant_angle:.4f}rad ({math.degrees(cant_angle):.2f}°)")
    print(f"スラック: {slack}mm")

if __name__ == "__main__":
    extract_building_limit_data()
    analyze_calculation_formulas()