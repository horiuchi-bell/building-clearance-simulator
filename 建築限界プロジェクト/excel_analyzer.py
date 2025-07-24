#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OIRANシミュレーターExcelファイル解析ツール
建築限界モデルの設定データを抽出・解析
"""

import openpyxl
import pandas as pd
import numpy as np
import json
from typing import Dict, List, Tuple, Any

class OIRANExcelAnalyzer:
    def __init__(self, excel_path: str):
        """
        Excelファイル解析クラスの初期化
        
        Args:
            excel_path (str): Excelファイルのパス
        """
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet = None
        self.data = {}
        
        try:
            self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
            print(f"✅ Excelファイルを正常に読み込みました: {excel_path}")
        except FileNotFoundError:
            print(f"❌ Excelファイルが見つかりません: {excel_path}")
        except Exception as e:
            print(f"❌ Excelファイル読み込みエラー: {e}")
    
    def list_sheets(self) -> List[str]:
        """
        ワークブック内の全シート名を取得
        
        Returns:
            List[str]: シート名のリスト
        """
        if self.workbook:
            sheets = self.workbook.sheetnames
            print(f"📋 検出されたシート一覧 ({len(sheets)}件):")
            for i, sheet in enumerate(sheets, 1):
                print(f"  {i}. {sheet}")
            return sheets
        return []
    
    def find_target_sheet(self, target_name: str = "限界余裕測定図 片線") -> str:
        """
        目的のシートを検索
        
        Args:
            target_name (str): 検索するシート名
            
        Returns:
            str: 見つかったシート名（完全一致または部分一致）
        """
        if not self.workbook:
            return None
            
        sheets = self.workbook.sheetnames
        
        # 完全一致を優先
        if target_name in sheets:
            print(f"🎯 目的のシートを発見: {target_name}")
            return target_name
        
        # 部分一致検索
        for sheet in sheets:
            if target_name in sheet or sheet in target_name:
                print(f"🎯 類似シートを発見: {sheet}")
                return sheet
        
        print(f"❌ 目的のシート「{target_name}」が見つかりません")
        return None
    
    def analyze_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """
        指定シートのデータを解析
        
        Args:
            sheet_name (str): 解析対象のシート名
            
        Returns:
            Dict[str, Any]: 解析結果
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return {}
        
        self.worksheet = self.workbook[sheet_name]
        print(f"📊 シート「{sheet_name}」を解析中...")
        
        analysis_result = {
            "sheet_name": sheet_name,
            "dimensions": self._get_sheet_dimensions(),
            "parameters": self._extract_parameters(),
            "graph_data": self._analyze_graph_area(),
            "formulas": self._extract_formulas(),
            "cant_relations": self._find_cant_relations(),
            "curve_radius_relations": self._find_curve_radius_relations()
        }
        
        return analysis_result
    
    def _get_sheet_dimensions(self) -> Dict[str, int]:
        """シートの寸法情報を取得"""
        if not self.worksheet:
            return {}
        
        max_row = self.worksheet.max_row
        max_col = self.worksheet.max_column
        
        print(f"📏 シート寸法: {max_row}行 × {max_col}列")
        
        return {
            "max_row": max_row,
            "max_column": max_col
        }
    
    def _extract_parameters(self) -> Dict[str, Any]:
        """重要なパラメータを抽出"""
        parameters = {}
        
        # セル範囲を走査してパラメータを検索
        for row in range(1, min(50, self.worksheet.max_row + 1)):  # 最初の50行をスキャン
            for col in range(1, min(20, self.worksheet.max_column + 1)):  # 最初の20列をスキャン
                cell = self.worksheet.cell(row=row, column=col)
                
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    
                    # カント関連の検索
                    if "カント" in cell_value:
                        adjacent_value = self._get_adjacent_numeric_value(row, col)
                        if adjacent_value is not None:
                            parameters["cant"] = {
                                "position": f"{openpyxl.utils.get_column_letter(col)}{row}",
                                "value": adjacent_value
                            }
                    
                    # 曲線半径関連の検索
                    if "曲線半径" in cell_value or "半径" in cell_value:
                        adjacent_value = self._get_adjacent_numeric_value(row, col)
                        if adjacent_value is not None:
                            parameters["curve_radius"] = {
                                "position": f"{openpyxl.utils.get_column_letter(col)}{row}",
                                "value": adjacent_value
                            }
                    
                    # 測定値関連の検索
                    if "測定" in cell_value:
                        adjacent_value = self._get_adjacent_numeric_value(row, col)
                        if adjacent_value is not None:
                            key = cell_value.replace(":", "").replace("：", "")
                            parameters[key] = {
                                "position": f"{openpyxl.utils.get_column_letter(col)}{row}",
                                "value": adjacent_value
                            }
        
        print(f"🔍 抽出されたパラメータ: {len(parameters)}件")
        for key, value in parameters.items():
            print(f"  - {key}: {value}")
        
        return parameters
    
    def _get_adjacent_numeric_value(self, row: int, col: int) -> float:
        """隣接するセルから数値を取得"""
        # 右隣、下隣、右下のセルをチェック
        for dr, dc in [(0, 1), (1, 0), (1, 1), (0, 2), (2, 0)]:
            try:
                adjacent_cell = self.worksheet.cell(row=row+dr, column=col+dc)
                if adjacent_cell.value and isinstance(adjacent_cell.value, (int, float)):
                    return float(adjacent_cell.value)
                elif adjacent_cell.value and isinstance(adjacent_cell.value, str):
                    # 数値文字列の場合の処理
                    try:
                        return float(str(adjacent_cell.value).replace("mm", "").replace("m", "").strip())
                    except ValueError:
                        continue
            except:
                continue
        return None
    
    def _analyze_graph_area(self) -> Dict[str, Any]:
        """グラフエリアの解析"""
        graph_data = {
            "coordinates": [],
            "boundaries": None,
            "scale_info": {}
        }
        
        # グラフの座標データを探索
        numeric_data = []
        for row in range(1, self.worksheet.max_row + 1):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    numeric_data.append({
                        "row": row,
                        "col": col,
                        "value": cell.value,
                        "position": f"{openpyxl.utils.get_column_letter(col)}{row}"
                    })
        
        print(f"📈 グラフエリア: {len(numeric_data)}個の数値データを検出")
        
        # X,Y座標らしきデータのペアを検索
        coordinates = []
        for i in range(len(numeric_data)):
            for j in range(i+1, len(numeric_data)):
                data1 = numeric_data[i]
                data2 = numeric_data[j]
                
                # 隣接するセルまたは同じ行/列の場合、座標ペアの可能性
                if (abs(data1["row"] - data2["row"]) <= 1 and 
                    abs(data1["col"] - data2["col"]) <= 2):
                    coordinates.append({
                        "x": data1["value"],
                        "y": data2["value"],
                        "x_pos": data1["position"],
                        "y_pos": data2["position"]
                    })
        
        graph_data["coordinates"] = coordinates[:20]  # 最初の20個のペアを保存
        return graph_data
    
    def _extract_formulas(self) -> List[Dict[str, str]]:
        """数式を抽出"""
        formulas = []
        
        for row in range(1, self.worksheet.max_row + 1):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                
                # 数式が設定されているセルを検索
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    formula_info = {
                        "position": f"{openpyxl.utils.get_column_letter(col)}{row}",
                        "formula": str(cell.value) if cell.value else "",
                        "result": cell.value
                    }
                    formulas.append(formula_info)
        
        print(f"🧮 数式: {len(formulas)}個を検出")
        return formulas
    
    def _find_cant_relations(self) -> List[Dict[str, Any]]:
        """カント関連の計算式・関係性を検索"""
        cant_relations = []
        
        for row in range(1, self.worksheet.max_row + 1):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).lower()
                    if "cant" in cell_value or "カント" in cell_value:
                        # 周辺のセルを調査
                        relation_info = {
                            "position": f"{openpyxl.utils.get_column_letter(col)}{row}",
                            "content": cell.value,
                            "adjacent_values": self._get_surrounding_values(row, col)
                        }
                        cant_relations.append(relation_info)
        
        return cant_relations
    
    def _find_curve_radius_relations(self) -> List[Dict[str, Any]]:
        """曲線半径関連の計算式・関係性を検索"""
        radius_relations = []
        
        for row in range(1, self.worksheet.max_row + 1):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).lower()
                    if "半径" in cell_value or "radius" in cell_value or "曲線" in cell_value:
                        relation_info = {
                            "position": f"{openpyxl.utils.get_column_letter(col)}{row}",
                            "content": cell.value,
                            "adjacent_values": self._get_surrounding_values(row, col)
                        }
                        radius_relations.append(relation_info)
        
        return radius_relations
    
    def _get_surrounding_values(self, row: int, col: int, radius: int = 2) -> Dict[str, Any]:
        """指定セル周辺の値を取得"""
        surrounding = {}
        
        for dr in range(-radius, radius + 1):
            for dc in range(-radius, radius + 1):
                if dr == 0 and dc == 0:
                    continue
                
                try:
                    target_row = row + dr
                    target_col = col + dc
                    
                    if target_row > 0 and target_col > 0:
                        cell = self.worksheet.cell(row=target_row, column=target_col)
                        if cell.value:
                            pos_key = f"{openpyxl.utils.get_column_letter(target_col)}{target_row}"
                            surrounding[pos_key] = cell.value
                except:
                    continue
        
        return surrounding
    
    def save_analysis_result(self, result: Dict[str, Any], output_path: str = "analysis_result.json"):
        """解析結果をJSONファイルに保存"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2, default=str)
            print(f"💾 解析結果を保存しました: {output_path}")
        except Exception as e:
            print(f"❌ ファイル保存エラー: {e}")

def main():
    """メイン実行関数"""
    excel_path = "OIRANシュミレーター（修正）20231215.xlsx"
    target_sheet = "限界余裕測定図 片線"
    
    print("🚀 OIRANシミュレーター Excel解析を開始...")
    
    # 解析器を初期化
    analyzer = OIRANExcelAnalyzer(excel_path)
    
    if analyzer.workbook is None:
        print("❌ Excelファイルの読み込みに失敗しました")
        return
    
    # シート一覧を表示
    sheets = analyzer.list_sheets()
    
    # 目的のシートを検索
    target_sheet_name = analyzer.find_target_sheet(target_sheet)
    
    if target_sheet_name:
        # シートを解析
        result = analyzer.analyze_sheet(target_sheet_name)
        
        # 結果を保存
        analyzer.save_analysis_result(result, "building_clearance_analysis.json")
        
        print("✅ 解析完了！")
        return result
    else:
        print("❌ 目的のシートが見つかりませんでした")
        return None

if __name__ == "__main__":
    main()