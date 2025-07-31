#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OIRANシミュレーターExcelファイル解析ツール
指定されたシートとセルから必要な情報を抽出します
"""

import openpyxl
import json
from pathlib import Path

def analyze_oiran_simulator(file_path):
    """OIRANシミュレーターのExcelファイルを解析"""
    
    # Excelファイルを開く
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 利用可能なシート名を確認
    print("利用可能なシート:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    results = {}
    
    # 1. 「必要離れ計算シート片線」シートを探す
    target_sheet_names = ["必要離れ計算シート片線", "必要離れ計算シート", "計算シート片線"]
    sheet = None
    
    for sheet_name in target_sheet_names:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            results['sheet_name'] = sheet_name
            print(f"\n'{sheet_name}'シートが見つかりました")
            break
    
    if sheet:
        # B15～B20セルの必要離れ値を確認
        print("\nB15～B20セルの値:")
        necessary_clearance = {}
        for row in range(15, 21):
            cell_ref = f"B{row}"
            value = sheet[cell_ref].value
            label_cell = f"A{row}"
            label = sheet[label_cell].value
            
            # D列の値も確認（実際の計算結果が入っている可能性）
            d_cell_ref = f"D{row}"
            d_value = sheet[d_cell_ref].value
            
            # E列、G列も確認（交流・非電化区間の値）
            e_value = sheet[f"E{row}"].value
            g_value = sheet[f"G{row}"].value
            
            necessary_clearance[cell_ref] = {
                'label': label,
                'B_value': value,
                'D_value': d_value,
                'E_value': e_value,
                'G_value': g_value
            }
            print(f"  {cell_ref} ({label}): B={value}, D={d_value}, E={e_value}, G={g_value}")
        
        results['necessary_clearance_B15_B20'] = necessary_clearance
        
        # A12セルの曲線半径による拡幅の値を確認
        a12_value = sheet["A12"].value
        b12_value = sheet["B12"].value
        print(f"\nA12セル: {a12_value}")
        print(f"B12セル: {b12_value}")
        
        results['curve_expansion'] = {
            'A12': a12_value,
            'B12': b12_value
        }
        
        # 周辺のセルも確認（文脈を理解するため）
        print("\nA10～A14の内容:")
        context_cells = {}
        for row in range(10, 15):
            cell_ref = f"A{row}"
            value = sheet[cell_ref].value
            b_value = sheet[f"B{row}"].value
            context_cells[cell_ref] = {
                'A_column': value,
                'B_column': b_value
            }
            print(f"  A{row}: {value} | B{row}: {b_value}")
        
        results['context_A10_A14'] = context_cells
    
    # 2. カントによる傾き補正量の計算式を探す
    print("\n\nカントに関連するシートを探索中...")
    cant_related_sheets = []
    cant_formulas = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # シート名にカント関連のキーワードが含まれているか確認
        if any(keyword in sheet_name for keyword in ["カント", "傾き", "傾斜", "cant", "CANT"]):
            cant_related_sheets.append(sheet_name)
        
        # 各シートでカント関連のセルを探す
        print(f"\n'{sheet_name}'シートでカント関連の内容を検索中...")
        for row in range(1, min(50, sheet.max_row + 1)):
            for col in range(1, min(10, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if any(keyword in str(cell.value) for keyword in ["カント", "傾き補正", "傾斜補正"]):
                        cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        
                        # 隣接セルの値も取得
                        adjacent_value = None
                        if col < sheet.max_column:
                            adjacent_value = sheet.cell(row=row, column=col+1).value
                        
                        cant_formulas[f"{sheet_name}_{cell_ref}"] = {
                            'sheet': sheet_name,
                            'cell': cell_ref,
                            'value': cell.value,
                            'adjacent_value': adjacent_value
                        }
                        print(f"  {cell_ref}: {cell.value} (隣接セル: {adjacent_value})")
    
    results['cant_related_sheets'] = cant_related_sheets
    results['cant_formulas'] = cant_formulas
    
    # 3. 数式を含むセルを探す（data_only=Falseで再読み込み）
    print("\n\n数式を確認するため、ファイルを再読み込み中...")
    wb_formula = openpyxl.load_workbook(file_path, data_only=False)
    
    formulas = {}
    for sheet_name in wb_formula.sheetnames[:5]:  # 最初の5シートのみ確認
        sheet = wb_formula[sheet_name]
        print(f"\n'{sheet_name}'シートの数式を確認中...")
        
        for row in range(1, min(30, sheet.max_row + 1)):
            for col in range(1, min(10, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    formulas[f"{sheet_name}_{cell_ref}"] = {
                        'sheet': sheet_name,
                        'cell': cell_ref,
                        'formula': cell.value
                    }
                    print(f"  {cell_ref}: {cell.value}")
    
    results['formulas'] = formulas
    
    # 結果をJSONファイルに保存
    output_file = Path(file_path).parent / "oiran_analysis_results.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n\n解析結果を {output_file} に保存しました")
    
    return results


if __name__ == "__main__":
    file_path = "/home/tems_kaihatu/建築限界プロジェクト/OIRANシュミレーター（修正）20231215.xlsx"
    
    try:
        results = analyze_oiran_simulator(file_path)
        
        print("\n" + "="*50)
        print("解析結果サマリー:")
        print("="*50)
        
        if 'necessary_clearance_B15_B20' in results:
            print("\n必要離れ値（B15～B20）:")
            for cell, data in results['necessary_clearance_B15_B20'].items():
                print(f"  {cell}: B={data['B_value']}, D={data['D_value']} ({data['label']})")
        
        if 'curve_expansion' in results:
            print(f"\n曲線半径による拡幅:")
            print(f"  A12: {results['curve_expansion']['A12']}")
            print(f"  B12: {results['curve_expansion']['B12']}")
        
        if 'cant_formulas' in results and results['cant_formulas']:
            print(f"\nカント関連の項目: {len(results['cant_formulas'])}件")
            
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()