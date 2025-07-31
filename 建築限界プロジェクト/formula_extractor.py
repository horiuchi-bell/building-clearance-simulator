#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
建築限界プロジェクト - 数式詳細抽出ツール
三角関数（カント計算）と建築限界座標変換の解析
"""

import openpyxl
import json
import re
from typing import Dict, List, Any

class FormulaExtractor:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path, data_only=False)
        
    def extract_trigonometric_formulas(self, sheet_name: str) -> List[Dict[str, Any]]:
        """三角関数を含む数式を詳細抽出"""
        ws = self.workbook[sheet_name]
        trig_formulas = []
        
        print(f"🔍 {sheet_name} から三角関数を抽出中...")
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    
                    if hasattr(cell, 'data_type') and cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        
                        # 三角関数を含む数式を検索
                        if any(func in formula.upper() for func in ["SIN", "COS", "TAN", "ATAN", "ASIN", "ACOS"]):
                            pos = f"{openpyxl.utils.get_column_letter(col)}{row}"
                            
                            # 数式を詳細解析
                            analysis = self.analyze_formula(formula)
                            
                            trig_info = {
                                "position": pos,
                                "row": row,
                                "col": col,
                                "formula": formula,
                                "functions_used": analysis["functions"],
                                "references": analysis["references"],
                                "constants": analysis["constants"],
                                "likely_purpose": self.identify_formula_purpose(formula)
                            }
                            
                            trig_formulas.append(trig_info)
                            
                except Exception as e:
                    continue
        
        print(f"✅ {len(trig_formulas)} 個の三角関数数式を抽出")
        return trig_formulas
    
    def analyze_formula(self, formula: str) -> Dict[str, List[str]]:
        """数式の詳細解析"""
        analysis = {
            "functions": [],
            "references": [],
            "constants": []
        }
        
        # 三角関数を検索
        trig_functions = re.findall(r'(SIN|COS|TAN|ATAN|ASIN|ACOS)\(', formula.upper())
        analysis["functions"] = list(set(trig_functions))
        
        # セル参照を検索（例: A1, $A$1, Sheet!A1）
        cell_refs = re.findall(r"(?:'[^']*'!)?(?:\$?[A-Z]+\$?\d+)", formula)
        analysis["references"] = list(set(cell_refs))
        
        # 数値定数を検索
        constants = re.findall(r'\b\d+\.?\d*\b', formula)
        analysis["constants"] = list(set(constants))
        
        return analysis
    
    def identify_formula_purpose(self, formula: str) -> str:
        """数式の目的を推定"""
        formula_upper = formula.upper()
        
        # カント関連の可能性を判定
        if "SIN" in formula_upper or "COS" in formula_upper:
            if any(keyword in formula_upper for keyword in ["PI", "180", "RADIANS"]):
                return "cant_angle_calculation"
            else:
                return "trigonometric_transformation"
        
        # 角度計算
        elif "ATAN" in formula_upper or "ASIN" in formula_upper or "ACOS" in formula_upper:
            return "angle_calculation"
        
        # その他
        else:
            return "general_trigonometric"
    
    def extract_cant_curve_inputs(self, sheet_name: str) -> Dict[str, Any]:
        """カント・曲線半径の入力値を検索"""
        ws = self.workbook[sheet_name]
        inputs = {
            "cant_inputs": [],
            "curve_radius_inputs": [],
            "related_parameters": []
        }
        
        print(f"🔍 {sheet_name} からカント・曲線半径入力を検索...")
        
        for row in range(1, min(30, ws.max_row + 1)):  # 上位30行をスキャン
            for col in range(1, min(15, ws.max_column + 1)):  # 左側15列をスキャン
                try:
                    cell = ws.cell(row=row, column=col)
                    
                    if cell.value and isinstance(cell.value, str):
                        cell_value = str(cell.value).strip()
                        pos = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        
                        # カント関連
                        if "カント" in cell_value:
                            adjacent_value = self.get_adjacent_numeric(ws, row, col)
                            inputs["cant_inputs"].append({
                                "position": pos,
                                "label": cell_value,
                                "value": adjacent_value,
                                "adjacent_position": self.get_adjacent_position(row, col, adjacent_value)
                            })
                        
                        # 曲線半径関連
                        elif "曲線半径" in cell_value or "半径" in cell_value:
                            adjacent_value = self.get_adjacent_numeric(ws, row, col)
                            inputs["curve_radius_inputs"].append({
                                "position": pos,
                                "label": cell_value,
                                "value": adjacent_value,
                                "adjacent_position": self.get_adjacent_position(row, col, adjacent_value)
                            })
                        
                        # その他の重要パラメータ
                        elif any(keyword in cell_value for keyword in ["速度", "軌間", "電源", "高さ"]):
                            adjacent_value = self.get_adjacent_numeric(ws, row, col)
                            if adjacent_value is not None:
                                inputs["related_parameters"].append({
                                    "position": pos,
                                    "label": cell_value,
                                    "value": adjacent_value
                                })
                
                except Exception as e:
                    continue
        
        return inputs
    
    def get_adjacent_numeric(self, ws, row: int, col: int) -> float:
        """隣接セルから数値を取得"""
        for dr, dc in [(0, 1), (0, 2), (1, 0), (1, 1), (0, -1), (-1, 0)]:
            try:
                target_cell = ws.cell(row=row+dr, column=col+dc)
                if isinstance(target_cell.value, (int, float)):
                    return float(target_cell.value)
                elif isinstance(target_cell.value, str):
                    # 数値文字列の処理
                    try:
                        return float(re.sub(r'[^\d.-]', '', target_cell.value))
                    except:
                        continue
            except:
                continue
        return None
    
    def get_adjacent_position(self, row: int, col: int, value: float) -> str:
        """数値が見つかった位置を返す"""
        if value is None:
            return ""
        
        # この実装では簡略化して隣の位置を返す
        return f"{openpyxl.utils.get_column_letter(col+1)}{row}"
    
    def extract_coordinate_transformation_formulas(self, sheet_name: str) -> List[Dict[str, Any]]:
        """座標変換数式を抽出（建築限界の形状変換）"""
        ws = self.workbook[sheet_name]
        coord_formulas = []
        
        print(f"🔍 {sheet_name} から座標変換数式を抽出...")
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    
                    if hasattr(cell, 'data_type') and cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        
                        # 座標変換らしい数式を検索
                        if (any(func in formula.upper() for func in ["SIN", "COS"]) and 
                            any(op in formula for op in ["+", "-", "*"])):
                            
                            pos = f"{openpyxl.utils.get_column_letter(col)}{row}"
                            coord_info = {
                                "position": pos,
                                "formula": formula,
                                "analysis": self.analyze_formula(formula),
                                "coordinate_type": self.identify_coordinate_type(row, col, formula)
                            }
                            coord_formulas.append(coord_info)
                            
                except Exception as e:
                    continue
        
        print(f"✅ {len(coord_formulas)} 個の座標変換数式を抽出")
        return coord_formulas
    
    def identify_coordinate_type(self, row: int, col: int, formula: str) -> str:
        """座標タイプを推定（X座標 or Y座標）"""
        # 列位置や数式の内容から推定
        if col <= 10:  # 左側の列
            return "likely_x_coordinate"
        elif col > 10:  # 右側の列
            return "likely_y_coordinate"
        else:
            return "unknown_coordinate"
    
    def analyze_all_formula_relationships(self) -> Dict[str, Any]:
        """全数式の関係性を解析"""
        key_sheets = [
            "表示データ計算シート　片線",
            "表示データ　片線",
            "建築限界数値データ　片線"
        ]
        
        results = {
            "trigonometric_analysis": {},
            "input_parameters": {},
            "coordinate_transformations": {},
            "summary": {}
        }
        
        for sheet_name in key_sheets:
            if sheet_name in self.workbook.sheetnames:
                print(f"\n📊 {sheet_name} を解析中...")
                
                # 三角関数解析
                trig_formulas = self.extract_trigonometric_formulas(sheet_name)
                results["trigonometric_analysis"][sheet_name] = trig_formulas
                
                # 入力パラメータ解析
                inputs = self.extract_cant_curve_inputs(sheet_name)
                results["input_parameters"][sheet_name] = inputs
                
                # 座標変換解析
                coord_formulas = self.extract_coordinate_transformation_formulas(sheet_name)
                results["coordinate_transformations"][sheet_name] = coord_formulas
                
                # サマリー
                results["summary"][sheet_name] = {
                    "trigonometric_count": len(trig_formulas),
                    "cant_inputs": len(inputs["cant_inputs"]),
                    "curve_inputs": len(inputs["curve_radius_inputs"]),
                    "coordinate_formulas": len(coord_formulas)
                }
        
        return results
    
    def save_formula_analysis(self, results: Dict[str, Any], filename: str = "formula_analysis.json"):
        """数式解析結果を保存"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2, default=str)
            print(f"💾 数式解析結果を保存: {filename}")
        except Exception as e:
            print(f"❌ 保存エラー: {e}")

def main():
    """メイン実行"""
    print("🧮 建築限界 数式詳細解析を開始...")
    
    extractor = FormulaExtractor("OIRANシュミレーター（修正）20231215.xlsx")
    results = extractor.analyze_all_formula_relationships()
    
    # 結果サマリーを表示
    print("\n📋 数式解析サマリー:")
    for sheet_name, summary in results["summary"].items():
        print(f"\n🔸 {sheet_name}:")
        print(f"  - 三角関数: {summary['trigonometric_count']}個")
        print(f"  - カント入力: {summary['cant_inputs']}個")
        print(f"  - 曲線半径入力: {summary['curve_inputs']}個")
        print(f"  - 座標変換数式: {summary['coordinate_formulas']}個")
    
    # カント・曲線半径の具体的な値を表示
    print("\n🎯 検出された入力パラメータ:")
    for sheet_name, inputs in results["input_parameters"].items():
        if inputs["cant_inputs"] or inputs["curve_radius_inputs"]:
            print(f"\n📍 {sheet_name}:")
            for cant_input in inputs["cant_inputs"]:
                print(f"  カント: {cant_input['value']} ({cant_input['position']})")
            for curve_input in inputs["curve_radius_inputs"]:
                print(f"  曲線半径: {curve_input['value']} ({curve_input['position']})")
    
    # 結果を保存
    extractor.save_formula_analysis(results)
    
    print("\n✅ 数式詳細解析完了！")
    return results

if __name__ == "__main__":
    main()