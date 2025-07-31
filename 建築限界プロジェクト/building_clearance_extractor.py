#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
建築限界プロジェクト - 基本形状データ抽出ツール
建築限界の基本座標データとカント・曲線半径変換式の抽出
"""

import openpyxl
import numpy as np
import json
from typing import Dict, List, Tuple, Any

class BuildingClearanceExtractor:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
    def extract_base_clearance_shape(self, sheet_name: str = "建築限界数値データ　片線") -> List[Tuple[float, float]]:
        """基本建築限界形状の座標データを抽出"""
        ws = self.workbook[sheet_name]
        coordinates = []
        
        print(f"🔍 {sheet_name} から基本形状データを抽出中...")
        
        # 数値データを全て収集
        numeric_data = {}
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        pos = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        numeric_data[pos] = {
                            "value": float(cell.value),
                            "row": row,
                            "col": col
                        }
                except:
                    continue
        
        print(f"📊 {len(numeric_data)} 個の数値データを検出")
        
        # 座標ペアを形成（隣接する数値を X,Y座標として扱う）
        coordinate_pairs = []
        processed_positions = set()
        
        for pos, data in numeric_data.items():
            if pos in processed_positions:
                continue
                
            row, col = data["row"], data["col"]
            x_value = data["value"]
            
            # 右隣のセルをチェック（Y座標として）
            y_pos = f"{openpyxl.utils.get_column_letter(col+1)}{row}"
            if y_pos in numeric_data:
                y_value = numeric_data[y_pos]["value"]
                
                # 建築限界として妥当な範囲かチェック
                if self.is_valid_clearance_coordinate(x_value, y_value):
                    coordinate_pairs.append({
                        "x": x_value,
                        "y": y_value,
                        "x_pos": pos,
                        "y_pos": y_pos,
                        "row": row
                    })
                    processed_positions.add(pos)
                    processed_positions.add(y_pos)
        
        # Y座標順にソート（下から上へ）
        coordinate_pairs.sort(key=lambda p: p["y"])
        
        print(f"✅ {len(coordinate_pairs)} 個の座標ペアを抽出")
        
        # タプルのリストとして返す
        return [(pair["x"], pair["y"]) for pair in coordinate_pairs]
    
    def is_valid_clearance_coordinate(self, x: float, y: float) -> bool:
        """建築限界座標として妥当かチェック"""
        # 一般的な建築限界の範囲（mm単位）
        x_range = (-4000, 6000)  # X座標範囲
        y_range = (-500, 5000)   # Y座標範囲
        
        return (x_range[0] <= x <= x_range[1] and 
                y_range[0] <= y <= y_range[1])
    
    def extract_cant_transformation_logic(self, sheet_name: str = "表示データ計算シート　片線") -> Dict[str, Any]:
        """カント変換ロジックの抽出"""
        ws = self.workbook[sheet_name]
        
        # カント入力値を取得
        cant_value = ws.cell(row=1, column=4).value  # D1
        curve_radius = ws.cell(row=1, column=5).value  # E1
        
        print(f"📐 カント値: {cant_value}, 曲線半径: {curve_radius}")
        
        # 主要な変換数式を抽出
        transformation_formulas = {}
        
        # 数式サンプルを取得（最初の数行）
        for row in range(2, min(10, ws.max_row + 1)):
            for col in range(1, min(30, ws.max_column + 1)):
                try:
                    cell = ws.cell(row=row, column=col)
                    if hasattr(cell, 'data_type') and cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        if any(func in formula.upper() for func in ["SIN", "COS", "TAN", "ATAN"]):
                            pos = f"{openpyxl.utils.get_column_letter(col)}{row}"
                            transformation_formulas[pos] = {
                                "formula": formula,
                                "result": cell.value if hasattr(cell, 'value') else None
                            }
                except:
                    continue
        
        print(f"🧮 {len(transformation_formulas)} 個の変換数式を抽出")
        
        return {
            "cant_value": cant_value,
            "curve_radius": curve_radius,
            "formulas": transformation_formulas
        }
    
    def analyze_cant_angle_calculation(self, cant_value: float) -> Dict[str, float]:
        """カント角度計算の解析"""
        # Excelの数式: =IF(D2<0,(ATAN(D2/1067)+2*PI()),ATAN(D2/1067))
        # 1067は軌間の可能性（1067mm = 標準軌）
        
        gauge = 1067  # 軌間（mm）
        
        if cant_value < 0:
            angle_rad = np.arctan(cant_value / gauge) + 2 * np.pi
        else:
            angle_rad = np.arctan(cant_value / gauge)
        
        angle_deg = np.degrees(angle_rad)
        
        return {
            "cant_mm": cant_value,
            "gauge_mm": gauge,
            "angle_rad": angle_rad,
            "angle_deg": angle_deg,
            "sin_value": np.sin(angle_rad),
            "cos_value": np.cos(angle_rad)
        }
    
    def transform_coordinates_with_cant(self, coordinates: List[Tuple[float, float]], 
                                      cant_value: float) -> List[Tuple[float, float]]:
        """カント値による座標変換"""
        if cant_value == 0:
            return coordinates
        
        angle_info = self.analyze_cant_angle_calculation(cant_value)
        sin_angle = angle_info["sin_value"]
        cos_angle = angle_info["cos_value"]
        
        transformed_coords = []
        
        for x, y in coordinates:
            # 2D回転変換行列を適用
            # x' = x * cos(θ) - y * sin(θ)
            # y' = x * sin(θ) + y * cos(θ)
            
            new_x = x * cos_angle - y * sin_angle
            new_y = x * sin_angle + y * cos_angle
            
            transformed_coords.append((new_x, new_y))
        
        return transformed_coords
    
    def calculate_curve_widening(self, coordinates: List[Tuple[float, float]], 
                               curve_radius: float) -> List[Tuple[float, float]]:
        """曲線半径による拡幅計算"""
        if curve_radius == 0 or curve_radius > 10000:  # 直線とみなす
            return coordinates
        
        # 曲線での拡幅計算（簡略化）
        # 実際のロジックはExcelから詳細に抽出する必要がある
        widening_factor = 1000.0 / curve_radius  # 仮の計算式
        
        widened_coords = []
        for x, y in coordinates:
            # X座標のみ拡幅（レール側を広げる）
            if x > 0:  # 右側
                new_x = x + widening_factor
            elif x < 0:  # 左側
                new_x = x - widening_factor
            else:
                new_x = x
            
            widened_coords.append((new_x, y))
        
        return widened_coords
    
    def generate_transformed_clearance(self, cant_value: float = 0, 
                                     curve_radius: float = 0) -> Dict[str, Any]:
        """変換された建築限界を生成"""
        print(f"🚀 建築限界変換を実行: カント={cant_value}mm, 曲線半径={curve_radius}m")
        
        # 基本形状を取得
        base_coordinates = self.extract_base_clearance_shape()
        
        if not base_coordinates:
            print("❌ 基本形状データが見つかりません")
            return {}
        
        # カント変換を適用
        cant_transformed = self.transform_coordinates_with_cant(base_coordinates, cant_value)
        
        # 曲線拡幅を適用
        final_coordinates = self.calculate_curve_widening(cant_transformed, curve_radius)
        
        # ランク判定（簡略化）
        rank = self.determine_clearance_rank(cant_value, curve_radius)
        
        return {
            "input_parameters": {
                "cant_mm": cant_value,
                "curve_radius_m": curve_radius
            },
            "base_coordinates": base_coordinates,
            "cant_transformed": cant_transformed,
            "final_coordinates": final_coordinates,
            "rank": rank,
            "coordinate_count": len(final_coordinates)
        }
    
    def determine_clearance_rank(self, cant_value: float, curve_radius: float) -> str:
        """建築限界ランクを判定"""
        # Excelの複雑な判定ロジックを簡略化
        if abs(cant_value) > 100:  # 100mm以上のカント
            if abs(cant_value) > 200:
                return "A"  # 支障あり
            else:
                return "B"  # 支障あり
        elif abs(cant_value) > 50:
            return "D"
        elif abs(cant_value) > 0:
            return "E"
        else:
            return "E"  # カントなし
    
    def save_clearance_data(self, data: Dict[str, Any], filename: str = "clearance_data.json"):
        """建築限界データを保存"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
            print(f"💾 建築限界データを保存: {filename}")
        except Exception as e:
            print(f"❌ 保存エラー: {e}")

def main():
    """メイン実行"""
    print("🏗️ 建築限界基本形状データ抽出を開始...")
    
    extractor = BuildingClearanceExtractor("OIRANシュミレーター（修正）20231215.xlsx")
    
    # 基本形状を抽出
    base_shape = extractor.extract_base_clearance_shape()
    print(f"📐 基本形状: {len(base_shape)} 点")
    
    # カント変換ロジックを抽出
    cant_logic = extractor.extract_cant_transformation_logic()
    
    # テストケース: 複数のカント・曲線半径の組み合わせ
    test_cases = [
        {"cant": 0, "radius": 0},      # 傾きなし
        {"cant": 140, "radius": 300},  # 参考画像の条件
        {"cant": 100, "radius": 600},  # 中間値
        {"cant": -80, "radius": 1000}, # 負のカント
    ]
    
    results = {}
    for i, case in enumerate(test_cases):
        print(f"\n🧪 テストケース {i+1}: カント={case['cant']}mm, 曲線半径={case['radius']}m")
        result = extractor.generate_transformed_clearance(case["cant"], case["radius"])
        results[f"case_{i+1}"] = result
    
    # 全結果を保存
    final_data = {
        "base_shape": base_shape,
        "cant_logic": cant_logic,
        "test_results": results
    }
    
    extractor.save_clearance_data(final_data, "building_clearance_data.json")
    
    print("\n✅ 建築限界基本形状データ抽出完了！")
    return final_data

if __name__ == "__main__":
    main()