#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
建築限界プロジェクト - 詳細Excel解析ツール
カント・曲線半径の計算式と建築限界データの詳細抽出
"""

import openpyxl
import pandas as pd
import numpy as np
import json
from typing import Dict, List, Tuple, Any

class DetailedExcelAnalyzer:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path, data_only=False)  # 数式も取得
        self.key_sheets = [
            "限界余裕測定図 片線",
            "表示データ　片線", 
            "表示データ計算シート　片線",
            "建築限界数値データ　片線",
            "限界判定計算シート　片線"
        ]
        
    def analyze_all_key_sheets(self) -> Dict[str, Any]:
        """重要なシートを全て解析"""
        results = {}
        
        for sheet_name in self.key_sheets:
            if sheet_name in self.workbook.sheetnames:
                print(f"\n📊 解析中: {sheet_name}")
                results[sheet_name] = self.analyze_sheet_detailed(sheet_name)
            else:
                print(f"⚠️  シートが見つかりません: {sheet_name}")
        
        return results
    
    def analyze_sheet_detailed(self, sheet_name: str) -> Dict[str, Any]:
        """詳細なシート解析"""
        ws = self.workbook[sheet_name]
        
        result = {
            "basic_info": {
                "name": sheet_name,
                "dimensions": f"{ws.max_row}行 × {ws.max_column}列"
            },
            "input_parameters": self.extract_input_parameters(ws),
            "calculated_values": self.extract_calculated_values(ws),
            "formulas": self.extract_all_formulas(ws),
            "coordinate_data": self.extract_coordinate_data(ws),
            "cant_curve_relations": self.extract_cant_curve_relations(ws)
        }
        
        return result
    
    def extract_input_parameters(self, ws) -> Dict[str, Any]:
        """入力パラメータの抽出"""
        parameters = {}
        
        # よくある入力欄の位置をチェック
        common_input_areas = [
            (range(1, 20), range(1, 10)),  # 左上エリア
            (range(1, 15), range(10, 20)), # 右上エリア
        ]
        
        for row_range, col_range in common_input_areas:
            for row in row_range:
                for col in col_range:
                    try:
                        cell = ws.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str):
                            value_str = str(cell.value).strip()
                            
                            # カント関連
                            if "カント" in value_str:
                                param_info = self.get_parameter_info(ws, row, col, "カント")
                                if param_info:
                                    parameters["cant"] = param_info
                            
                            # 曲線半径関連
                            elif "曲線半径" in value_str or "半径" in value_str:
                                param_info = self.get_parameter_info(ws, row, col, "曲線半径")
                                if param_info:
                                    parameters["curve_radius"] = param_info
                            
                            # その他の重要パラメータ
                            elif any(keyword in value_str for keyword in ["測定", "離れ", "高さ", "ランク"]):
                                param_key = value_str.replace(":", "").replace("：", "").strip()
                                param_info = self.get_parameter_info(ws, row, col, param_key)
                                if param_info:
                                    parameters[param_key] = param_info
                    except:
                        continue
        
        return parameters
    
    def get_parameter_info(self, ws, row: int, col: int, param_name: str) -> Dict[str, Any]:
        """パラメータの詳細情報を取得"""
        # 隣接するセルから値を検索
        for dr, dc in [(0, 1), (0, 2), (1, 0), (1, 1), (0, -1), (-1, 0)]:
            try:
                target_row = row + dr
                target_col = col + dc
                target_cell = ws.cell(row=target_row, column=target_col)
                
                if target_cell.value is not None:
                    pos = f"{openpyxl.utils.get_column_letter(target_col)}{target_row}"
                    
                    # 数値の場合
                    if isinstance(target_cell.value, (int, float)):
                        return {
                            "name": param_name,
                            "position": pos,
                            "value": target_cell.value,
                            "formula": target_cell.value if hasattr(target_cell, 'value') else None,
                            "type": "numeric"
                        }
                    
                    # 文字列の場合（数値が含まれているかチェック）
                    elif isinstance(target_cell.value, str):
                        try:
                            numeric_part = self.extract_numeric_from_string(target_cell.value)
                            if numeric_part is not None:
                                return {
                                    "name": param_name,
                                    "position": pos,
                                    "value": numeric_part,
                                    "original_text": target_cell.value,
                                    "type": "text_with_numeric"
                                }
                        except:
                            continue
            except:
                continue
        
        return None
    
    def extract_numeric_from_string(self, text: str) -> float:
        """文字列から数値を抽出"""
        import re
        
        # 数値パターンを検索
        patterns = [
            r'(\d+\.?\d*)',  # 基本的な数値
            r'(\d+)',        # 整数
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text)
            if matches:
                try:
                    return float(matches[0])
                except:
                    continue
        
        return None
    
    def extract_calculated_values(self, ws) -> Dict[str, Any]:
        """計算値の抽出"""
        calculated = {}
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    
                    # 数式セルを検索
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        pos = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        calculated[pos] = {
                            "formula": str(cell.value) if cell.value else "",
                            "result": cell.value,
                            "position": pos
                        }
                except:
                    continue
        
        return calculated
    
    def extract_all_formulas(self, ws) -> List[Dict[str, Any]]:
        """全ての数式を抽出"""
        formulas = []
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        pos = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        formula_info = {
                            "position": pos,
                            "formula": str(cell.value) if cell.value else "",
                            "coordinates": {"row": row, "col": col}
                        }
                        
                        # 数式の内容を解析
                        if cell.value:
                            formula_text = str(cell.value)
                            if "SIN" in formula_text or "COS" in formula_text:
                                formula_info["type"] = "trigonometric"
                            elif "SQRT" in formula_text:
                                formula_info["type"] = "mathematical"
                            elif any(ref in formula_text for ref in ["$", ":"]):
                                formula_info["type"] = "reference"
                            else:
                                formula_info["type"] = "arithmetic"
                        
                        formulas.append(formula_info)
                except:
                    continue
        
        return formulas
    
    def extract_coordinate_data(self, ws) -> Dict[str, Any]:
        """座標データの抽出（建築限界の形状データ）"""
        coordinates = {
            "x_values": [],
            "y_values": [],
            "coordinate_pairs": []
        }
        
        # 数値データを全て収集
        numeric_cells = []
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        numeric_cells.append({
                            "row": row,
                            "col": col,
                            "value": cell.value,
                            "position": f"{openpyxl.utils.get_column_letter(col)}{row}"
                        })
                except:
                    continue
        
        # 座標データっぽいペアを検索
        for i, cell1 in enumerate(numeric_cells):
            for cell2 in numeric_cells[i+1:]:
                # 隣接するセルまたは近い位置にある数値ペア
                if (abs(cell1["row"] - cell2["row"]) <= 2 and 
                    abs(cell1["col"] - cell2["col"]) <= 3):
                    
                    # X,Y座標として妥当な範囲の値かチェック
                    val1, val2 = cell1["value"], cell2["value"]
                    
                    # 建築限界の一般的な範囲（mm単位）
                    if (-3000 <= val1 <= 6000 and -1000 <= val2 <= 5000):
                        coordinates["coordinate_pairs"].append({
                            "x": val1,
                            "y": val2,
                            "x_pos": cell1["position"],
                            "y_pos": cell2["position"]
                        })
        
        return coordinates
    
    def extract_cant_curve_relations(self, ws) -> Dict[str, Any]:
        """カント・曲線半径の関係性を抽出"""
        relations = {
            "cant_formulas": [],
            "curve_formulas": [],
            "transformation_formulas": []
        }
        
        # 数式を解析してカント・曲線半径に関連するものを抽出
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    
                    if hasattr(cell, 'data_type') and cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        pos = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        
                        # 三角関数（傾き計算）を含む数式
                        if any(func in formula.upper() for func in ["SIN", "COS", "TAN", "ATAN"]):
                            relations["transformation_formulas"].append({
                                "position": pos,
                                "formula": formula,
                                "type": "trigonometric",
                                "purpose": "likely_cant_calculation"
                            })
                        
                        # 曲線半径に関連する計算（距離、拡幅など）
                        elif any(keyword in formula.upper() for keyword in ["SQRT", "POWER", "ABS"]):
                            relations["curve_formulas"].append({
                                "position": pos,
                                "formula": formula,
                                "type": "mathematical",
                                "purpose": "likely_curve_calculation"
                            })
                except:
                    continue
        
        return relations
    
    def save_detailed_analysis(self, results: Dict[str, Any], filename: str = "detailed_analysis.json"):
        """詳細解析結果を保存"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2, default=str)
            print(f"💾 詳細解析結果を保存: {filename}")
        except Exception as e:
            print(f"❌ 保存エラー: {e}")

def main():
    """メイン実行"""
    print("🔍 建築限界 詳細Excel解析を開始...")
    
    analyzer = DetailedExcelAnalyzer("OIRANシュミレーター（修正）20231215.xlsx")
    results = analyzer.analyze_all_key_sheets()
    
    # 結果をサマリー表示
    print("\n📋 解析結果サマリー:")
    for sheet_name, data in results.items():
        print(f"\n🔸 {sheet_name}:")
        print(f"  - 入力パラメータ: {len(data.get('input_parameters', {}))}個")
        print(f"  - 計算値: {len(data.get('calculated_values', {}))}個")
        print(f"  - 数式: {len(data.get('formulas', []))}個")
        print(f"  - 座標データ: {len(data.get('coordinate_data', {}).get('coordinate_pairs', []))}個")
        
        # 重要な発見をハイライト
        formulas = data.get('formulas', [])
        trig_formulas = [f for f in formulas if f.get('type') == 'trigonometric']
        if trig_formulas:
            print(f"  ⭐ 三角関数: {len(trig_formulas)}個 (カント計算の可能性)")
    
    # 結果を保存
    analyzer.save_detailed_analysis(results)
    
    print("\n✅ 詳細解析完了！")
    return results

if __name__ == "__main__":
    main()