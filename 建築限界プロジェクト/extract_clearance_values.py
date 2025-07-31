#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OIRANシミュレーターから必要離れ値を抽出する専用スクリプト
"""

import openpyxl
from pathlib import Path

def extract_clearance_values():
    """必要離れ計算シート片線から必要な値を抽出"""
    
    # Excelファイルを開く
    file_path = "/home/tems_kaihatu/建築限界プロジェクト/OIRANシュミレーター（修正）20231215.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 「必要離れ計算シート 片線」シートを取得
    if "必要離れ計算シート 片線" in wb.sheetnames:
        sheet = wb["必要離れ計算シート 片線"]
        
        print("=== 必要離れ計算シート 片線の解析結果 ===\n")
        
        # 1. B15～B20セルの必要離れ値（計算式と実際の値）
        print("【1. B15～B20セルの必要離れ値】")
        print("高さ範囲別の必要離れ計算式:")
        for row in range(15, 21):
            a_label = sheet[f"A{row}"].value
            b_formula = sheet[f"B{row}"].value
            c_condition = sheet[f"C{row}"].value
            d_value = sheet[f"D{row}"].value
            e_value = sheet[f"E{row}"].value
            g_value = sheet[f"G{row}"].value
            
            print(f"\n行{row}:")
            if a_label:
                print(f"  ラベル(A{row}): {a_label}")
            print(f"  計算式(B{row}): {b_formula}")
            print(f"  条件(C{row}): {c_condition}")
            print(f"  直流区間値(D{row}): {d_value}")
            print(f"  交流区間値(E{row}): {e_value}")
            print(f"  交流区間値(G{row}): {g_value}")
        
        # 2. A12セルの曲線半径による拡幅の値
        print("\n\n【2. 曲線半径による拡幅】")
        a12_value = sheet["A12"].value
        b12_value = sheet["B12"].value if "B12" in sheet else None
        print(f"  A12セル: {a12_value}")
        print(f"  B12セル: {b12_value}")
        
        # 周辺のセルも確認
        print("\n  A10～A14の内容:")
        for row in range(10, 15):
            a_value = sheet[f"A{row}"].value
            b_value = sheet[f"B{row}"].value if f"B{row}" in sheet else None
            print(f"    A{row}: {a_value} | B{row}: {b_value}")
        
        # 3. カントによる傾き補正量
        print("\n\n【3. カントによる傾き補正量】")
        
        # D3, D4のカント関連情報
        d3_value = sheet["D3"].value
        d4_value = sheet["D4"].value
        print(f"  D3（カントラベル）: {d3_value}")
        print(f"  D4（カント値）: {d4_value}")
        
        # E8のカント関連の計算式を確認
        e8_value = sheet["E8"].value
        print(f"\n  E8（カント関連計算）: {e8_value}")
        
        # カント関連の数式を確認するため、data_only=Falseで再読み込み
        wb_formula = openpyxl.load_workbook(file_path, data_only=False)
        sheet_formula = wb_formula["必要離れ計算シート 片線"]
        
        print("\n  カント関連の計算式:")
        # B15～B20の計算式
        for row in range(15, 21):
            b_formula = sheet_formula[f"B{row}"].value
            if b_formula and isinstance(b_formula, str) and b_formula.startswith('='):
                print(f"    B{row}: {b_formula}")
        
        # E8の計算式（カント補正）
        e8_formula = sheet_formula["E8"].value
        if e8_formula and isinstance(e8_formula, str) and e8_formula.startswith('='):
            print(f"\n  E8のカント補正式: {e8_formula}")
            print("  ※カント<=0の場合: 1900+w+920*SIN(-1*ｔ)")
            print("  ※カント>0の場合: 1900+w+3156*SIN(-1*ｔ)")
        
        # 4. その他の重要なパラメータ
        print("\n\n【4. その他の重要なパラメータ】")
        
        # A4, B4, C4, D4の値（x, y, yd, カント）
        a4_value = sheet["A4"].value
        b4_value = sheet["B4"].value
        c4_value = sheet["C4"].value
        print(f"  A4（x座標）: {a4_value}")
        print(f"  B4（y座標）: {b4_value}")
        print(f"  C4（yd高さ）: {c4_value}")
        
        # A8, B8の計算結果
        a8_value = sheet["A8"].value
        b8_value = sheet["B8"].value
        d8_value = sheet["D8"].value
        print(f"\n  A8（x方向の離れ）: {a8_value}")
        print(f"  B8（y方向の離れ）: {b8_value}")
        print(f"  D8（必要離れ合計）: {d8_value}")
        
        # 5. 変数の定義を探す
        print("\n\n【5. 使用されている変数】")
        print("  w: 曲線半径による拡幅量")
        print("  yd: 測定高さ")
        print("  y: y座標")
        print("  x: x座標")
        print("  ｔ(t): カントによる傾き角度")
        print("  カント: カント量（mm）")
        
    else:
        print("「必要離れ計算シート 片線」シートが見つかりません")
    
    wb.close()


if __name__ == "__main__":
    extract_clearance_values()