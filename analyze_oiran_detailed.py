#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OIRANシュミレーターExcelファイルの詳細分析スクリプト
"""

import openpyxl
import pandas as pd
import json

def analyze_excel_detailed():
    """Excelファイルの詳細分析"""
    excel_file = "OIRANシュミレーター（修正）20231215.xlsx"
    
    try:
        # 数式を含む分析のため data_only=False で読み込み
        workbook = openpyxl.load_workbook(excel_file, data_only=False)
        
        # メインシート「限界余裕測定図 片線」の詳細分析
        target_sheet = "限界余裕測定図 片線"
        
        if target_sheet in workbook.sheetnames:
            print(f"=== {target_sheet} 詳細分析 ===")
            sheet = workbook[target_sheet]
            
            # 入力セル周辺の数式を詳細確認
            print("\n=== 入力セル周辺の数式詳細 ===")
            for row in range(9, 20):
                for col in range(1, 10):  # A-I列
                    cell = sheet.cell(row=row, column=col)
                    col_letter = openpyxl.utils.get_column_letter(col)
                    
                    if cell.value is not None or (hasattr(cell, 'formula') and cell.formula):
                        formula_info = ""
                        if hasattr(cell, 'formula') and cell.formula:
                            formula_info = f" [式: {cell.formula}]"
                        print(f"{col_letter}{row}: {cell.value}{formula_info}")
            
            # 建築限界データの計算に関連する可能性のある数式を探す
            print("\n=== 建築限界関連の数式セル ===")
            formula_cells = []
            for row in range(1, min(50, sheet.max_row + 1)):
                for col in range(1, min(20, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)
                    if hasattr(cell, 'formula') and cell.formula:
                        col_letter = openpyxl.utils.get_column_letter(col)
                        formula_cells.append({
                            'cell': f"{col_letter}{row}",
                            'formula': cell.formula,
                            'value': cell.value
                        })
            
            # 数式を含むセルを表示
            for i, formula_info in enumerate(formula_cells):
                print(f"{formula_info['cell']}: {formula_info['formula']}")
                if i >= 30:  # 最初の30個だけ表示
                    print(f"...他 {len(formula_cells) - 30} 個の数式セル")
                    break
            
            # 他のシートとの参照関係を確認
            print("\n=== 他シートへの参照 ===")
            for formula_info in formula_cells:
                if "!" in formula_info['formula']:
                    print(f"{formula_info['cell']}: {formula_info['formula']}")
            
            # 数値データを含む列を広範囲で探索
            print("\n=== 数値データ列の探索 ===")
            for col in range(8, min(30, sheet.max_column + 1)):  # H列以降
                col_letter = openpyxl.utils.get_column_letter(col)
                numeric_data = []
                for row in range(1, min(50, sheet.max_row + 1)):
                    cell = sheet.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)) and cell.value != 0:
                        numeric_data.append((row, cell.value))
                
                if len(numeric_data) > 3:  # 数値データが多い列を表示
                    print(f"\n{col_letter}列の数値データ:")
                    for row, value in numeric_data[:15]:
                        print(f"  {col_letter}{row}: {value}")
                    if len(numeric_data) > 15:
                        print(f"  ...他 {len(numeric_data) - 15} 行")
        
        # 関連シートの確認
        related_sheets = ["建築限界数値データ　片線", "表示データ　片線", "限界余裕計算シート 片線"]
        
        for sheet_name in related_sheets:
            if sheet_name in workbook.sheetnames:
                print(f"\n=== {sheet_name} シート概要 ===")
                sheet = workbook[sheet_name]
                
                # 数値データの分布を確認
                numeric_cells = []
                for row in range(1, min(50, sheet.max_row + 1)):
                    for col in range(1, min(20, sheet.max_column + 1)):
                        cell = sheet.cell(row=row, column=col)
                        if isinstance(cell.value, (int, float)) and cell.value != 0:
                            col_letter = openpyxl.utils.get_column_letter(col)
                            numeric_cells.append(f"{col_letter}{row}: {cell.value}")
                
                print(f"数値データ（最初の20個）:")
                for i, cell_info in enumerate(numeric_cells[:20]):
                    print(f"  {cell_info}")
                if len(numeric_cells) > 20:
                    print(f"  ...他 {len(numeric_cells) - 20} 個")
        
        workbook.close()
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_excel_detailed()