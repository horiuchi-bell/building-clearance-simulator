#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OIRANシュミレーターExcelファイルの数式分析スクリプト
"""

import openpyxl
import pandas as pd

def analyze_excel_formulas():
    """Excelファイルの数式を詳細分析"""
    excel_file = "OIRANシュミレーター（修正）20231215.xlsx"
    
    try:
        # 数式を含む分析のため data_only=False で読み込み
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        
        # メインシート分析
        target_sheet = "限界余裕測定図 片線"
        
        if target_sheet in wb.sheetnames:
            print(f"=== {target_sheet} 数式分析 ===")
            ws = wb[target_sheet]
            
            # 全セルをスキャンして数式を探す
            print("\n=== 全セルスキャン（数式） ===")
            formulas_found = []
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    col_letter = openpyxl.utils.get_column_letter(col)
                    
                    # 数式かどうかを確認
                    if cell.data_type == 'f':  # 数式セル
                        formulas_found.append({
                            'cell': f"{col_letter}{row}",
                            'formula': cell.value,
                            'coordinate': f"{col_letter}{row}"
                        })
            
            print(f"数式セル数: {len(formulas_found)}")
            
            # 数式を表示
            for i, formula_info in enumerate(formulas_found):
                print(f"{formula_info['cell']}: {formula_info['formula']}")
                if i >= 50:  # 最初の50個だけ表示
                    print(f"...他 {len(formulas_found) - 50} 個の数式セル")
                    break
            
            # 重要な範囲の詳細確認
            print("\n=== 重要セル範囲の詳細確認 ===")
            important_ranges = [
                (10, 18, 1, 10),  # 入力セル周辺
                (1, 25, 8, 17),   # グラフ部分推定
                (25, 43, 1, 17),  # 下部データ
            ]
            
            for row_start, row_end, col_start, col_end in important_ranges:
                print(f"\n範囲: {openpyxl.utils.get_column_letter(col_start)}{row_start}:{openpyxl.utils.get_column_letter(col_end)}{row_end}")
                
                for row in range(row_start, min(row_end + 1, ws.max_row + 1)):
                    for col in range(col_start, min(col_end + 1, ws.max_column + 1)):
                        cell = ws.cell(row=row, column=col)
                        col_letter = openpyxl.utils.get_column_letter(col)
                        
                        if cell.value is not None:
                            info = f"{col_letter}{row}: {cell.value}"
                            if cell.data_type == 'f':
                                info += f" [数式]"
                            print(info)
        
        # 関連シートの詳細確認
        related_sheets = ["建築限界数値データ　片線", "表示データ　片線", "限界余裕計算シート 片線"]
        
        for sheet_name in related_sheets:
            if sheet_name in wb.sheetnames:
                print(f"\n=== {sheet_name} 詳細データ ===")
                ws = wb[sheet_name]
                
                # 全データを確認
                data_cells = []
                for row in range(1, min(30, ws.max_row + 1)):
                    for col in range(1, min(15, ws.max_column + 1)):
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None:
                            col_letter = openpyxl.utils.get_column_letter(col)
                            data_type = "数式" if cell.data_type == 'f' else "値"
                            data_cells.append(f"{col_letter}{row}: {cell.value} [{data_type}]")
                
                # 最初の30個を表示
                for i, cell_info in enumerate(data_cells[:30]):
                    print(f"  {cell_info}")
                if len(data_cells) > 30:
                    print(f"  ...他 {len(data_cells) - 30} 個")
        
        wb.close()
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_excel_formulas()